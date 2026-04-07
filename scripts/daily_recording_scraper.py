#!/usr/bin/env python3
"""
Daily Recording Scraper for Jump Contact

Scrapes the previous day's call recordings from Twilio via the TaskRouter
chain (call_sid -> task -> conference -> recording) and writes them to a
per-day output file.

Designed to run once per day, ideally after midnight MST so the prior day
is fully closed. Follows the verified F20/F21 pattern from the build spec.

USAGE:
    python daily_recording_scraper.py                  # yesterday
    python daily_recording_scraper.py 2026-04-06       # specific date

OUTPUT:
    recordings_YYYY-MM-DD.xlsx in current directory

ENV VARS REQUIRED:
    TWILIO_ACCOUNT_SID
    TWILIO_AUTH_TOKEN
    TWILIO_WORKSPACE_SID
"""

import json
import os
import sys
import subprocess
import time
from datetime import datetime, timedelta, timezone

# ============================================================
# CONFIG
# ============================================================
ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', 'ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN')
WORKSPACE = os.environ.get('TWILIO_WORKSPACE_SID', 'WSxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')

if not AUTH_TOKEN:
    sys.exit("ERROR: Set TWILIO_AUTH_TOKEN environment variable")

BASE = f"https://api.twilio.com/2010-04-01/Accounts/{ACCOUNT_SID}"
BASE_TR = f"https://taskrouter.twilio.com/v1/Workspaces/{WORKSPACE}"

# ============================================================
# HTTP HELPER (uses curl to avoid extra dependencies)
# ============================================================
def api_get(url):
    """GET against Twilio API with Basic auth. Returns parsed JSON."""
    out = subprocess.run(
        ['curl', '-s', '-u', f'{ACCOUNT_SID}:{AUTH_TOKEN}', url],
        capture_output=True, text=True, timeout=30
    )
    if out.returncode != 0:
        raise RuntimeError(f"curl failed: {out.stderr}")
    return json.loads(out.stdout)

# ============================================================
# DATE WINDOW
# ============================================================
if len(sys.argv) > 1:
    target_date = datetime.strptime(sys.argv[1], '%Y-%m-%d').replace(tzinfo=timezone.utc)
else:
    target_date = (datetime.now(timezone.utc) - timedelta(days=1)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )

day_start = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
day_end = day_start + timedelta(days=1)
date_str = day_start.strftime('%Y-%m-%d')

print(f"Scraping recordings for {date_str}")
print(f"  UTC window: {day_start.isoformat()} -> {day_end.isoformat()}")

# Calculate Minutes parameter for TaskRouter (events older than 14 days are dropped)
now = datetime.now(timezone.utc)
minutes_back = int((now - day_start).total_seconds() / 60) + 60  # +1hr buffer
if minutes_back > 20160:  # 14 days
    print(f"WARNING: Target date is beyond TaskRouter retention (14 days). Recent events only.")
    minutes_back = 20160

# ============================================================
# STEP 1: TaskRouter task.created events for the target day
# ============================================================
print(f"\nSTEP 1: Scanning TaskRouter task.created events...")
task_map = {}  # task_sid -> {call_sid, project, event_date}
url = f"{BASE_TR}/Events?Minutes={minutes_back}&EventType=task.created&PageSize=100"
page = 0

while url:
    page += 1
    data = api_get(url)
    events = data.get('events', [])

    for e in events:
        ed_date = e.get('event_date', '')
        if not ed_date:
            continue

        # Parse and filter to target day only
        try:
            edt = datetime.fromisoformat(ed_date.replace('Z', '+00:00'))
        except ValueError:
            continue
        if edt < day_start or edt >= day_end:
            continue

        ed = e.get('event_data', '{}')
        if isinstance(ed, str):
            try:
                ed = json.loads(ed)
            except json.JSONDecodeError:
                continue

        ta = ed.get('task_attributes', '')
        if isinstance(ta, str):
            try:
                ta = json.loads(ta)
            except json.JSONDecodeError:
                continue

        if isinstance(ta, dict):
            call_sid = ta.get('call_sid', '')
            project = ta.get('project', '')
            task_sid = ed.get('task_sid', '')
            if call_sid and task_sid:
                task_map[task_sid] = {
                    'call_sid': call_sid,
                    'project': project,
                    'event_date': ed_date,
                }

    nxt = data.get('meta', {}).get('next_page_url')
    url = nxt if nxt else None

    if page % 10 == 0:
        print(f"  Page {page}: {len(task_map)} tasks for {date_str}...")
    time.sleep(0.05)

print(f"  Done. {len(task_map)} tasks found for {date_str}")

# ============================================================
# STEP 2: Conferences for the target day
# ============================================================
print(f"\nSTEP 2: Listing conferences...")
task_to_conf = {}
url = (f"{BASE}/Conferences.json"
       f"?DateCreated%3E={date_str}"
       f"&DateCreated%3C={(day_start + timedelta(days=2)).strftime('%Y-%m-%d')}"
       f"&Status=completed&PageSize=1000")
page = 0
total_confs = 0

while url:
    page += 1
    data = api_get(url)
    confs = data.get('conferences', [])
    total_confs += len(confs)

    for c in confs:
        fname = c.get('friendly_name', '')
        if fname.startswith('WT'):
            task_to_conf[fname] = c['sid']

    nxt = data.get('next_page_uri')
    url = f"https://api.twilio.com{nxt}" if nxt else None
    time.sleep(0.1)

print(f"  Done. {total_confs} conferences, {len(task_to_conf)} task linkages")

# ============================================================
# STEP 3: Recordings for the target day
# ============================================================
print(f"\nSTEP 3: Listing recordings...")
conf_to_rec = {}
url = (f"{BASE}/Recordings.json"
       f"?DateCreated%3E={date_str}"
       f"&DateCreated%3C={(day_start + timedelta(days=2)).strftime('%Y-%m-%d')}"
       f"&PageSize=100")
page = 0
total_recs = 0

while url:
    page += 1
    data = api_get(url)
    recs = data.get('recordings', [])
    total_recs += len(recs)

    for r in recs:
        csid = r.get('conference_sid')
        if csid:
            conf_to_rec[csid] = {
                'rec_sid': r['sid'],
                'mp3': f"https://api.twilio.com{r['uri'].replace('.json', '.mp3')}",
                'dur': r['duration'],
                'date': r['date_created'],
            }

    nxt = data.get('next_page_uri')
    url = f"https://api.twilio.com{nxt}" if nxt else None
    time.sleep(0.05)

print(f"  Done. {total_recs} recordings, {len(conf_to_rec)} unique conferences")

# ============================================================
# STEP 4: Chain everything together
# ============================================================
print(f"\nSTEP 4: Building call -> recording map...")
results = []
matched = no_conf = no_rec = 0

for task_sid, info in task_map.items():
    call_sid = info['call_sid']
    project = info['project']

    conf_sid = task_to_conf.get(task_sid)
    if not conf_sid:
        no_conf += 1
        results.append({
            'date': date_str, 'client': project, 'call_sid': call_sid,
            'recording_sid': '', 'duration_sec': '', 'mp3_url': '',
            'status': 'queue_drop'
        })
        continue

    rec_info = conf_to_rec.get(conf_sid)
    if not rec_info:
        no_rec += 1
        results.append({
            'date': date_str, 'client': project, 'call_sid': call_sid,
            'recording_sid': '', 'duration_sec': '', 'mp3_url': '',
            'status': 'no_recording'
        })
        continue

    matched += 1
    results.append({
        'date': date_str,
        'client': project,
        'call_sid': call_sid,
        'recording_sid': rec_info['rec_sid'],
        'duration_sec': rec_info['dur'],
        'mp3_url': rec_info['mp3'],
        'status': 'matched'
    })

print(f"\n{'='*60}")
print(f"SUMMARY for {date_str}:")
print(f"  Total calls:        {len(results)}")
print(f"  Recordings matched: {matched}")
print(f"  Queue drops:        {no_conf}")
print(f"  No recording:       {no_rec}")
hit_rate = f"{matched/len(results)*100:.0f}%" if results else "n/a"
print(f"  Hit rate:           {hit_rate}")
print(f"{'='*60}")

# ============================================================
# STEP 5: Write XLSX
# ============================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"Recordings {date_str}"

    headers = ['Date', 'Client', 'Call SID', 'Recording SID', 'Duration (s)', 'MP3 URL', 'Status']
    hfill = PatternFill('solid', fgColor='1C1917')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    base_font = Font(name='Arial', size=10)
    mono = Font(name='Consolas', size=9)
    url_font = Font(name='Consolas', size=8, color='0369A1')
    bdr = Border(bottom=Side(style='thin', color='E7E5E4'))
    green_fill = PatternFill('solid', fgColor='F0FFF0')
    gfont = Font(color='15803D', bold=True, name='Arial', size=10)
    rfont = Font(color='B91C1C', name='Arial', size=10)
    xfont = Font(color='78716C', name='Arial', size=10)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill

    row = 2
    # Sort: matched first, then by client
    for r in sorted(results, key=lambda x: (x['status'] != 'matched', x['client'], x['call_sid'])):
        ws.cell(row=row, column=1, value=r['date']).font = base_font
        ws.cell(row=row, column=2, value=r['client']).font = base_font
        ws.cell(row=row, column=3, value=r['call_sid']).font = mono
        ws.cell(row=row, column=4, value=r['recording_sid']).font = mono
        ws.cell(row=row, column=5, value=r['duration_sec']).font = base_font
        ws.cell(row=row, column=6, value=r['mp3_url']).font = url_font

        status_label = {
            'matched': 'FOUND',
            'queue_drop': 'QUEUE DROP',
            'no_recording': 'NO REC'
        }.get(r['status'], r['status'])
        sc = ws.cell(row=row, column=7, value=status_label)
        sc.font = (gfont if r['status'] == 'matched'
                   else rfont if r['status'] == 'no_recording'
                   else xfont)

        if r['status'] == 'matched':
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = green_fill

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = bdr
        row += 1

    widths = {'A': 12, 'B': 35, 'C': 40, 'D': 40, 'E': 14, 'F': 100, 'G': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    output_file = f"recordings_{date_str}.xlsx"
    wb.save(output_file)
    print(f"\nWrote {output_file}")

except ImportError:
    # Fallback: CSV if openpyxl not installed
    import csv
    output_file = f"recordings_{date_str}.csv"
    with open(output_file, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['date', 'client', 'call_sid', 'recording_sid', 'duration_sec', 'mp3_url', 'status'])
        writer.writeheader()
        writer.writerows(results)
    print(f"\nWrote {output_file} (openpyxl not installed, used CSV)")

# Also write a JSON file for programmatic consumption
json_file = f"recordings_{date_str}.json"
with open(json_file, 'w') as f:
    json.dump({
        'date': date_str,
        'summary': {
            'total': len(results),
            'matched': matched,
            'queue_drops': no_conf,
            'no_recording': no_rec,
        },
        'recordings': results,
    }, f, indent=2)
print(f"Wrote {json_file}")
